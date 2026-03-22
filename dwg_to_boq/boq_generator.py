"""
BOQ Excel Generator
===================
Generates formatted Excel Bill of Quantities from classified MEP data.
Output structure matches the standard BOQ format:
  ITEM | DESCRIPTION | UNIT | QTY | UNIT/RATE | AMOUNT
organized by discipline sheets with element sections.
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .classifier import ClassifiedResult, BOQItem, DISCIPLINES

logger = logging.getLogger(__name__)

# Friendly sheet names
DISCIPLINE_SHEET_NAMES = {
    "ELECTRICAL": "ELECTRICAL",
    "PLUMBING": "PLUMBING",
    "HVAC": "HVAC",
    "FIRE_FIGHTING": "FIRE FIGHTING",
}


class BOQGenerator:
    """Generates Excel BOQ workbook from classified entities."""

    def __init__(self, config: dict):
        styles = config.get("excel_styles", {})
        self.header_fill = PatternFill(
            start_color=styles.get("header_fill", "1F4E79"),
            end_color=styles.get("header_fill", "1F4E79"),
            fill_type="solid",
        )
        self.header_font = Font(
            name="Calibri", size=11, bold=True,
            color=styles.get("header_font_color", "FFFFFF"),
        )
        self.element_fill = PatternFill(
            start_color=styles.get("element_fill", "D9E2F3"),
            end_color=styles.get("element_fill", "D9E2F3"),
            fill_type="solid",
        )
        self.summary_fill = PatternFill(
            start_color=styles.get("summary_fill", "FFF2CC"),
            end_color=styles.get("summary_fill", "FFF2CC"),
            fill_type="solid",
        )
        self.bold_font = Font(name="Calibri", size=11, bold=True)
        self.normal_font = Font(name="Calibri", size=10)
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate(
        self,
        result: ClassifiedResult,
        output_path: str,
        project_name: str = "",
    ):
        """Generate the Excel BOQ workbook."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        by_disc = result.by_discipline_and_element()

        discipline_totals: dict[str, int] = {}

        for discipline in DISCIPLINES:
            elements = by_disc.get(discipline, {})
            if not elements:
                continue

            sheet_name = DISCIPLINE_SHEET_NAMES.get(discipline, discipline)
            ws = wb.create_sheet(title=sheet_name)
            total_items = self._write_discipline_sheet(ws, discipline, elements, project_name)
            discipline_totals[sheet_name] = total_items

        # Write unclassified items if any
        if result.unclassified_blocks:
            ws = wb.create_sheet(title="UNCLASSIFIED")
            self._write_unclassified_sheet(ws, result.unclassified_blocks)

        # Write summary sheet
        if discipline_totals:
            ws_summary = wb.create_sheet(title="SUMMARY")
            self._write_summary_sheet(ws_summary, project_name, wb)
            # Move summary to the end
            wb.move_sheet("SUMMARY", offset=0)

        wb.save(output_path)
        logger.info(f"BOQ saved to: {output_path}")

    def _write_discipline_sheet(
        self,
        ws,
        discipline: str,
        elements: dict[str, list[BOQItem]],
        project_name: str,
    ) -> int:
        """Write a single discipline sheet. Returns total item count."""
        # Column widths
        col_widths = {"A": 8, "B": 55, "C": 8, "D": 10, "E": 15, "F": 18}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        title_cell = ws.cell(row=row, column=1)
        sheet_name = DISCIPLINE_SHEET_NAMES.get(discipline, discipline)
        title_cell.value = f"{project_name} - {sheet_name} BILL OF QUANTITIES" if project_name else f"{sheet_name} BILL OF QUANTITIES"
        title_cell.font = Font(name="Calibri", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        row += 2

        # Column headers
        headers = ["ITEM", "DESCRIPTION OF ITEMS", "UNIT", "QTY", "UNIT/RATE", "AMOUNT"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.thin_border
        row += 1

        element_num = 0
        total_items = 0
        element_summary_rows: list[tuple[str, int]] = []  # (element_name, summary_row)

        for element_name, items in sorted(elements.items()):
            element_num += 1

            # Element header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            elem_cell = ws.cell(row=row, column=1)
            elem_cell.value = f"ELEMENT {element_num}: {element_name.upper()}"
            elem_cell.font = self.bold_font
            elem_cell.fill = self.element_fill
            elem_cell.alignment = Alignment(horizontal="left")
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).fill = self.element_fill
            row += 1

            # Sort items by description
            items_sorted = sorted(items, key=lambda x: x.description)
            item_num = 0

            for item in items_sorted:
                item_num += 1
                total_items += 1

                item_label = f"{element_num}.{item_num}"

                # ITEM
                cell_a = ws.cell(row=row, column=1, value=item_label)
                cell_a.font = self.normal_font
                cell_a.alignment = Alignment(horizontal="center")
                cell_a.border = self.thin_border

                # DESCRIPTION
                cell_b = ws.cell(row=row, column=2, value=item.description)
                cell_b.font = self.normal_font
                cell_b.alignment = Alignment(wrap_text=True)
                cell_b.border = self.thin_border

                # UNIT
                cell_c = ws.cell(row=row, column=3, value=item.unit)
                cell_c.font = self.normal_font
                cell_c.alignment = Alignment(horizontal="center")
                cell_c.border = self.thin_border

                # QTY
                cell_d = ws.cell(row=row, column=4, value=item.quantity)
                cell_d.font = self.normal_font
                cell_d.alignment = Alignment(horizontal="center")
                cell_d.number_format = '#,##0'
                cell_d.border = self.thin_border

                # UNIT/RATE (empty - to be filled by estimator)
                cell_e = ws.cell(row=row, column=5)
                cell_e.font = self.normal_font
                cell_e.number_format = '#,##0.00'
                cell_e.border = self.thin_border

                # AMOUNT (formula: QTY * UNIT/RATE)
                cell_f = ws.cell(
                    row=row, column=6,
                    value=f"=D{row}*E{row}"
                )
                cell_f.font = self.normal_font
                cell_f.number_format = '#,##0.00'
                cell_f.border = self.thin_border

                row += 1

            # Element subtotal row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            subtotal_label = ws.cell(row=row, column=1)
            subtotal_label.value = f"CARRIED TO {sheet_name} BILL SUMMARY"
            subtotal_label.font = self.bold_font
            subtotal_label.alignment = Alignment(horizontal="right")
            subtotal_label.border = self.thin_border
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = self.thin_border

            # Sum formula for the element
            first_item_row = row - item_num
            subtotal_cell = ws.cell(
                row=row, column=6,
                value=f"=SUM(F{first_item_row}:F{row - 1})"
            )
            subtotal_cell.font = self.bold_font
            subtotal_cell.number_format = '#,##0.00'
            subtotal_cell.border = self.thin_border

            element_summary_rows.append((element_name, row))
            row += 2  # Blank row between elements

        # Bill summary at the bottom
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        summary_title = ws.cell(row=row, column=1)
        summary_title.value = f"{sheet_name} BILL SUMMARY"
        summary_title.font = Font(name="Calibri", size=12, bold=True)
        summary_title.fill = self.summary_fill
        summary_title.alignment = Alignment(horizontal="center")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.summary_fill
            ws.cell(row=row, column=col).border = self.thin_border
        row += 1

        for idx, (elem_name, summary_row) in enumerate(element_summary_rows, 1):
            cell_a = ws.cell(row=row, column=1, value=idx)
            cell_a.font = self.bold_font
            cell_a.alignment = Alignment(horizontal="center")
            cell_a.border = self.thin_border

            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            cell_b = ws.cell(row=row, column=2, value=elem_name)
            cell_b.font = self.bold_font
            cell_b.border = self.thin_border
            for col in range(3, 6):
                ws.cell(row=row, column=col).border = self.thin_border

            cell_f = ws.cell(row=row, column=6, value=f"=F{summary_row}")
            cell_f.font = self.bold_font
            cell_f.number_format = '#,##0.00'
            cell_f.border = self.thin_border
            row += 1

        # Grand total
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        gt_label = ws.cell(row=row, column=1)
        gt_label.value = f"GRAND TOTAL - {sheet_name}"
        gt_label.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
        gt_label.alignment = Alignment(horizontal="right")
        gt_label.border = self.thin_border
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = self.thin_border

        summary_start = row - len(element_summary_rows)
        gt_cell = ws.cell(
            row=row, column=6,
            value=f"=SUM(F{summary_start}:F{row - 1})"
        )
        gt_cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
        gt_cell.number_format = '#,##0.00'
        gt_cell.border = self.thin_border

        return total_items

    def _write_unclassified_sheet(self, ws, unclassified: dict[str, int]):
        """Write unclassified blocks for manual review."""
        col_widths = {"A": 8, "B": 50, "C": 8, "D": 10}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        title = ws.cell(row=row, column=1)
        title.value = "UNCLASSIFIED BLOCKS - REQUIRES MANUAL REVIEW"
        title.font = Font(name="Calibri", size=12, bold=True, color="FF0000")
        row += 2

        headers = ["#", "BLOCK NAME", "UNIT", "QTY"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        row += 1

        for idx, (block_name, count) in enumerate(
            sorted(unclassified.items(), key=lambda x: -x[1]), 1
        ):
            ws.cell(row=row, column=1, value=idx).border = self.thin_border
            ws.cell(row=row, column=2, value=block_name).border = self.thin_border
            ws.cell(row=row, column=3, value="NR").border = self.thin_border
            ws.cell(row=row, column=4, value=count).border = self.thin_border
            row += 1

    def _write_summary_sheet(self, ws, project_name: str, wb: Workbook):
        """Write the summary sheet that rolls up all discipline totals."""
        col_widths = {"A": 8, "B": 40, "C": 20}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        title = ws.cell(row=row, column=1)
        title.value = f"MEP COST SUMMARY - {project_name}" if project_name else "MEP COST SUMMARY"
        title.font = Font(name="Calibri", size=14, bold=True)
        title.alignment = Alignment(horizontal="center")
        row += 2

        headers = ["#", "DISCIPLINE", "AMOUNT"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.thin_border
        row += 1

        disc_num = 0
        first_amount_row = row
        for discipline in DISCIPLINES:
            sheet_name = DISCIPLINE_SHEET_NAMES.get(discipline, discipline)
            if sheet_name not in wb.sheetnames:
                continue

            disc_num += 1
            cell_a = ws.cell(row=row, column=1, value=disc_num)
            cell_a.font = self.bold_font
            cell_a.alignment = Alignment(horizontal="center")
            cell_a.border = self.thin_border

            cell_b = ws.cell(row=row, column=2, value=sheet_name)
            cell_b.font = self.bold_font
            cell_b.border = self.thin_border

            # Reference the grand total from each discipline sheet
            # Find the last row with data in the discipline sheet
            disc_ws = wb[sheet_name]
            grand_total_row = disc_ws.max_row
            cell_c = ws.cell(
                row=row, column=3,
                value=f"='{sheet_name}'!F{grand_total_row}"
            )
            cell_c.font = self.bold_font
            cell_c.number_format = '#,##0.00'
            cell_c.border = self.thin_border
            row += 1

        # Grand total
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        gt_label = ws.cell(row=row, column=1)
        gt_label.value = "TOTAL MEP COST"
        gt_label.font = Font(name="Calibri", size=12, bold=True, color="FF0000")
        gt_label.alignment = Alignment(horizontal="right")
        gt_label.fill = self.summary_fill
        ws.cell(row=row, column=2).fill = self.summary_fill
        gt_label.border = self.thin_border
        ws.cell(row=row, column=2).border = self.thin_border

        gt_cell = ws.cell(
            row=row, column=3,
            value=f"=SUM(C{first_amount_row}:C{row - 2})"
        )
        gt_cell.font = Font(name="Calibri", size=12, bold=True, color="FF0000")
        gt_cell.number_format = '#,##0.00'
        gt_cell.fill = self.summary_fill
        gt_cell.border = self.thin_border
